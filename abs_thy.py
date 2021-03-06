from selenium import webdriver
import time
import openpyxl
import os

# web sayfasi acilacak
# web sayfasindaki belirli noktaya awb nosu girilecek
# list dugmesine basilaca

for i in range(1, 3):
    os.chdir('/users/akin/desktop')
    awb_workbook = openpyxl.load_workbook('awb_numbers.xlsx')
    sheet1 = awb_workbook.get_sheet_by_name('Sheet1')

    awb_number = sheet1.cell(row=i, column=1).value
    
    adress = 'https://www.turkishcargo.com.tr/en/online-services/shipment-tracking'
    browser = webdriver.Chrome()
    browser.get(adress)

    awb_searchfield = browser.find_element_by_css_selector(
        '#serviceform > div:nth-child(6) > div.col-md-4.col-sm-4.col-xs-8 > div > div > input')
    time.sleep(3)
    awb_filled_searchfield = awb_searchfield.send_keys(awb_number)

    list_button = browser.find_element_by_css_selector(
        '#serviceform > div:nth-child(8) > div:nth-child(2) > button')
    list_button.click()
    time.sleep(3)

    last_movement = browser.find_element_by_css_selector(
        '#accordion > li > div.submenu > div:nth-child(1) > table > tbody > tr:nth-child(1) > td:nth-child(2)')
    last_movement = last_movement.text

    print(last_movement)

    os.chdir('/users/akin/desktop')
    workbook_results = openpyxl.load_workbook('thy_tracking.xlsx')
    sheet1 = workbook_results.get_sheet_by_name('Sheet')
    # last movement i al ve 2.columndan koymaya basla
    sheet1.cell(row=i, column=2).value = last_movement
    # AWB numarasini diger excelden alip buraya koy
    sheet1.cell(row=i, column=1).value = awb_number
            
    workbook_results.save('thy_tracking.xlsx')
